import pandas as pd
import time
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# === Mapping Org to Sheet ===
org_to_sheet = {
    "FLORIDA WOMAN CARE, LLC(70064)": "FL",
    "GENESIS OB/GYN(355998)": "AZ",
    "MID-ATLANTICWOMENSCARE, PLC(528579)": "MW",
    "MIDWEST CENTER FOR WOMEN'S HEALTHCARE(63366)": "IL",
    "MIDWEST CENTER FOR WOMEN'S HEALTHCARE(63366)": "IM",
    "NEW JERSEY PERINATAL ASSOCS(517429)": "NP",
    "PREMIER OBGYN OF MN(319136)": "MN",
    "UWH OF MICHIGAN, PLC(780558)": "MG",
    "OB/GYN Associates of Erie, PC(592003)" : "PE",
    "OB-GYN Associates of Erie Laboratory LLC(994204)" : "PE",
    "FWC REI LLC(411722)" : "FL",
    "FWC GYN ONCOLOGY LLC(411723)" : "FL",
    "FWC UROGYNECOLOGY LLC(411822)":"FL",
    "FWC PERINATAL LLC(411726)":"FL",
    "Square Care Medical Group, LLP(423268)": "NY",
    "UNIFIED WOMENS HEALTHCARE OF TEXAS(571405)": "TX",
    "UWH of North Carolina,LLP(463617)": ["NC","SC"]
}

def wait_for_file(filepath, timeout=20):
    print(f"Waiting for file: {filepath}")
    if os.path.basename(filepath).startswith('~$'):
        print(f"Error: The file {filepath} looks like a temporary Excel lock file.")
        sys.exit(1)
    start_time = time.time()
    while True:
        try:
            if os.path.exists(filepath):
                with open(filepath, 'rb'):
                    print(f"File ready: {filepath}")
                    return
        except Exception:
            pass
        if time.time() - start_time > timeout:
            print(f"Timeout: File not accessible after {timeout} seconds -> {filepath}")
            sys.exit(1)
        time.sleep(1)

def normalize_email(email):
    return str(email).strip().lower()

def normalize_status(status):
    return str(status).strip().upper()

def get_status_for_market(email, org_lookup, global_lookup):
    email = normalize_email(email)
    status = org_lookup.get(email)

    # First check org-specific
    if status:  
        status = normalize_status(status)   
    else:
        # Then fallback to global
        status = global_lookup.get(email)   
        if status:
            status = normalize_status(status)   
        else:
            return "Success - User not found"
 
    if status == "DEACTIVATED":
        return "Success - Deactivated"
    elif status == "EXPIRED INVITATION":
        return "Success - Expired Invitation"
    elif status == "PENDING INVITATION":
        return "Success - There is no option to deactivate for this status currently"
    elif status in ["ACTIVE", "LOCKED"]:
        return email
    else:
        return f"Unrecognized Status: {status}"

# === Main Execution ===
try:
    print("Starting Availity Email Processing...\n")
    time.sleep(1)

    if len(sys.argv) != 3:
        print("Usage: python SF_Union_Portals.py <salesforce_excel_path> <availity_excel_path>")
        sys.exit(1)

    sf_path = sys.argv[1]
    availity_path = sys.argv[2]

    wait_for_file(sf_path)
    wait_for_file(availity_path)

    try:
        availity_df = pd.read_excel(availity_path, engine='openpyxl')
    except Exception as e:
        print(f"Error reading Availity Excel: {str(e)}")
        sys.exit(1)

    required_columns = ['Email Address', 'Organization (Customer ID)', 'Status']
    for col in required_columns:
        if col not in availity_df.columns:
            print(f"Error: Column '{col}' not found in Availity Excel.")
            sys.exit(1)

    # Normalize
    availity_df['Email Address'] = availity_df['Email Address'].str.lower().str.strip()
    availity_df['Organization (Customer ID)'] = availity_df['Organization (Customer ID)'].str.strip().str.upper()
    availity_df['Status'] = availity_df['Status'].str.strip().str.upper()

    # Global status lookup
    global_status_lookup = availity_df.set_index('Email Address')['Status'].to_dict()

    # Email-to-active-orgs map
    email_to_active_orgs = {}
    for org_name in org_to_sheet:
        org_upper = org_name.strip().upper()
        market_df = availity_df[
            (availity_df['Organization (Customer ID)'] == org_upper) &
            (availity_df['Status'].isin(['ACTIVE', 'LOCKED']))
        ]
        for email in market_df['Email Address'].unique():
            email_to_active_orgs.setdefault(email, []).append(org_name)

    try:
        work_book = load_workbook(sf_path)
    except Exception as e:
        print(f"Error loading Salesforce Excel: {str(e)}")
        sys.exit(1)

    all_sheets = work_book.sheetnames

    # Build sheet_jobs to handle multiple sheets per org
    sheet_jobs = []
    for org_name, sheet_abbr in org_to_sheet.items():
        if isinstance(sheet_abbr, list):
            for abbr in sheet_abbr:
                sheet_jobs.append((org_name, abbr))
        else:
            sheet_jobs.append((org_name, sheet_abbr))
    if "MSO" in all_sheets:
        sheet_jobs.append(("MSO_ALL_ORGS", "MSO"))

    emailvalue_rows = []

    for org_name, sheet_abbr in sheet_jobs:
        if sheet_abbr not in all_sheets:
            print(f"Warning: Sheet '{sheet_abbr}' not found. Skipping.")
            continue

        try:
            print(f"Processing sheet: {sheet_abbr}")
            sf_df = pd.read_excel(sf_path, sheet_name=sheet_abbr, engine='openpyxl')
        except Exception as e:
            print(f"Error reading sheet '{sheet_abbr}': {str(e)}. Skipping.")
            continue

        if 'EMAIL' not in sf_df.columns:
            print(f"Warning: 'EMAIL' column not found in '{sheet_abbr}'. Skipping.")
            continue

        sf_df['EMAIL'] = sf_df['EMAIL'].str.lower().str.strip()

        if org_name == "MSO_ALL_ORGS":
            org_lookup = global_status_lookup
        else:
            org_upper = org_name.strip().upper()
            org_df = availity_df[availity_df['Organization (Customer ID)'] == org_upper]
            org_lookup = org_df.set_index('Email Address')['Status'].to_dict()

        try:
            sf_df['Availity'] = sf_df['EMAIL'].apply(lambda x: get_status_for_market(x, org_lookup, global_status_lookup))
        except Exception as e:
            print(f"Error applying status for sheet '{sheet_abbr}': {str(e)}. Skipping.")
            continue

        cols = [col for col in sf_df.columns if col != 'Availity'] + ['Availity']
        sf_df = sf_df[cols]

        try:
            work_sheet = work_book[sheet_abbr]
            work_sheet.delete_rows(1, work_sheet.max_row)
            for row in dataframe_to_rows(sf_df, index=False, header=True):
                work_sheet.append(row)
        except Exception as e:
            print(f"Error writing to sheet '{sheet_abbr}': {str(e)}")
            continue

        for idx, row in sf_df.iterrows():
            availity_value = row['Availity']
            if isinstance(availity_value, str) and re.match(r"[^@]+@[^@]+\.[^@]+", availity_value):
                email = row['EMAIL']
                active_orgs = ','.join(sorted(email_to_active_orgs.get(email, [])))
                emailvalue_rows.append([
                    email, idx + 2, sf_df.columns.get_loc('Availity') + 1, sheet_abbr, active_orgs
                ])

    try:
        work_book.save(sf_path)
        print("Excel sheets updated successfully.")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        sys.exit(1)

    # === Modified block to deduplicate before saving CSV ===
    if emailvalue_rows:
        try:
            timestamp = datetime.now().strftime("%d%m%Y")
            output_path = os.path.join(os.path.dirname(sf_path), f"{timestamp}_AvailityActiveEmails.csv")

            # Create DataFrame with all collected rows
            df = pd.DataFrame(emailvalue_rows, columns=['Email', 'Row', 'Column', 'Sheet', 'User Active Markets'])

            # Remove duplicate rows based on Email, Row, Column, and Sheet
            df = df.drop_duplicates(subset=['Email', 'Row', 'Column', 'Sheet'])

            # Write the deduplicated data to CSV
            df.to_csv(output_path, index=False)
            print(f"CSV saved: {output_path}")
        except Exception as e:
            print(f"Error writing CSV: {str(e)}")
            sys.exit(1)
    else:
        print("No active emails found. CSV not created.")

except Exception as e:
    print(f"Unexpected error: {str(e)}")
    sys.exit(1)
