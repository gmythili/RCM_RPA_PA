import pandas as pd
import time
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re

skip_markets = ["UL", "UM", "NE", "UG","CW","AG","OV","Other"] 
# Function to wait for a file to be accessible
def wait_for_file(filepath, timeout=20):
    print(f"Waiting for file: {filepath}")
    if os.path.basename(filepath).startswith('~$'):
        print(f"Error: The file {filepath} looks like a temporary Excel lock file (starts with '~$').")
        print("Please use the actual file, not the temp one.")
        sys.exit(1)

    start_time = time.time()
    while True:
        try:
            if os.path.exists(filepath):
                with open(filepath, 'rb'):
                    print(f"File ready: {filepath}")
                    return
        except Exception:
            pass
        if time.time() - start_time > timeout:
            print(f"Timeout: File not accessible after {timeout} seconds -> {filepath}")
            sys.exit(1)
        time.sleep(1)

# === Startup ===
print("Starting SF_Union_Portals processing script...\n")
time.sleep(2)

# === Get command line args ===
if len(sys.argv) != 3:
    print("Usage: python SF_Union_Portals.py <salesforce_excel_path> <uhc_csv_path>")
    sys.exit(1)

sf_path = sys.argv[1]
uhc_path = sys.argv[2]

wait_for_file(sf_path)
wait_for_file(uhc_path)

try:
    print("Reading UHC CSV...")
    try:
        uhc_df = pd.read_csv(uhc_path, encoding='utf-8')
    except UnicodeDecodeError:
        uhc_df = pd.read_csv(uhc_path, encoding='latin1')  # fallback

    uhc_df['Email Address'] = uhc_df['Email Address'].str.lower().str.strip()
    uhc_df['Market'] = uhc_df['Market'].str.strip()
    
    # Create a lookup dictionary for UHC status
    uhc_lookup = uhc_df.set_index('Email Address')['Status'].to_dict()
    
    # Create a mapping of active markets for each email
    active_markets_map = (
        uhc_df[uhc_df['Status'].str.lower() != 'inactive']
        .groupby('Email Address')['Market']
        .apply(lambda x: ','.join(sorted(set(x))))
        .to_dict()
    )

    # Function to check UHC status
    def check_uhc_status(email):
        email = str(email).strip().lower()
        status = uhc_lookup.get(email)
        if status is None:
            return 'Success - User not found'
        elif status.lower() == 'inactive':
            return 'Success - User found and already deactivated'
        else:
            return email
        
    # === Process Salesforce Excel file ===
    work_book = load_workbook(sf_path)
    all_sheets = work_book.sheetnames
    sheets_to_process = [s for s in all_sheets if s not in skip_markets]

    emailvalue_rows = []
    
    # Process each sheet
    for sheet_name in sheets_to_process:
        print(f"Processing sheet: {sheet_name}")
        sf_df = pd.read_excel(sf_path, sheet_name=sheet_name, engine='openpyxl')
        if 'EMAIL' not in sf_df.columns:
            print(f"'EMAIL' column not found in {sheet_name}. Skipping.")
            continue
        sf_df['EMAIL'] = sf_df['EMAIL'].str.lower().str.strip()
        sf_df['UHC'] = sf_df['EMAIL'].apply(check_uhc_status)

        work_sheet = work_book[sheet_name]
        work_sheet.delete_rows(1, work_sheet.max_row)
        for row in dataframe_to_rows(sf_df, index=False, header=True):
            work_sheet.append(row)

        # Update the 'UHC' column with the status
        for idx, val in sf_df['UHC'].items():
            if isinstance(val, str) and re.match(r"[^@]+@[^@]+\.[^@]+", val.strip()):
                email = val.strip()
                m_uhc = uhc_df.loc[uhc_df['Email Address'] == email, 'Market'].values
                m_uhc_val = m_uhc[0] if len(m_uhc) > 0 else ''
                all_markets = active_markets_map.get(email, '')
                emailvalue_rows.append([email, idx + 2, sf_df.columns.get_loc('UHC') + 1, sheet_name, m_uhc_val, all_markets])
    # Save the updated workbook
    work_book.save(sf_path)
    print("All sheets updated successfully.")

    ## === Create CSV with active UHC emails ===
    if emailvalue_rows:
        timestamp = datetime.now().strftime("%d%m%Y")
        output_filename = f"{timestamp}_UHCActive.csv"
        output_path = os.path.join(os.path.dirname(sf_path), output_filename)
        emailvalue_df = pd.DataFrame(emailvalue_rows, columns=['Email', 'Row', 'Column', 'Sheet', 'M_UHC', 'User Active Markets'])
        emailvalue_df.drop(columns=['M_UHC'], inplace=False)  # still keep M_UHC in CSV
        emailvalue_df.to_csv(output_path, index=False)
        print(f"CSV file saved: {output_path}")
    else:
        print("No active UHC emails found; no CSV created.")

except Exception as e:
    print("An error occurred:", str(e))
    sys.exit(1)