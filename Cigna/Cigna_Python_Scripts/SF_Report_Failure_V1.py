import sys
import os
import time
import re
from openpyxl import load_workbook

skip_markets = ["UL", "UM", "NE", "UG", "CW", "Other"]
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

# === File wait utility ===
def wait_for_file(filepath, timeout=20):
    """Wait until file is accessible and not a temporary Excel lock file."""
    print(f"Waiting for file: {filepath}")
    if os.path.basename(filepath).startswith("~$"):
        print(f"Error: {filepath} is a temporary Excel lock file. Provide actual file.")
        sys.exit(1)

    start_time = time.time()
    while True:
        try:
            if os.path.exists(filepath):
                with open(filepath, "rb"):
                    print(f"File ready: {filepath}")
                    return
        except Exception:
            pass
        if time.time() - start_time > timeout:
            print(f"Timeout: File not accessible after {timeout} seconds -> {filepath}")
            sys.exit(1)
        time.sleep(1)

# === Startup ===
print("Starting SF_Union_Portals Report Email Replacement Script...\n")
time.sleep(1)

# === Get args ===
if len(sys.argv) != 2:
    print("Usage: python SF_Report_Failure.py <salesforce_report_path>")
    sys.exit(1)

sf_report_path = sys.argv[1]
wait_for_file(sf_report_path)

try:
    # === Load workbook ===
    work_book = load_workbook(sf_report_path)
    all_sheets = work_book.sheetnames
    sheets_to_process = [s for s in all_sheets if s not in skip_markets]

    if not sheets_to_process:
        print("No valid sheets found to process. Exiting.")
        sys.exit(0)

    for sheet_name in sheets_to_process:
        print(f"Processing sheet: {sheet_name}")
        try:
            ws = work_book[sheet_name]

            # Skip if sheet empty
            if ws.max_row < 2 or ws.max_column < 1:
                print(f"Sheet {sheet_name} is empty. Skipping.")
                continue

            # Extract headers from first row
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            headers_upper = [h.upper().replace(" ", "_") for h in headers] 

            if "EMAIL" not in headers_upper:
                print(f"No 'EMAIL' column in {sheet_name}. Skipping.")
                continue

            email_col_idx = headers_upper.index("EMAIL") + 1  # 1-based index

            # Iterate rows (skip header row)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx, cell in enumerate(row, start=1):
                    if col_idx == email_col_idx:
                        continue  # skip EMAIL column

                    val = str(cell.value).strip() if cell.value else ""
                    if EMAIL_REGEX.match(val):
                        cell.value = "Failure - Action required"

        except Exception as sheet_err:
            print(f"Error processing sheet {sheet_name}: {sheet_err}")
            continue

    # Save updated workbook
    work_book.save(sf_report_path)
    print(f"\nReport updated successfully -> {sf_report_path}")

except Exception as e:
    print("An error occurred:", str(e))
    sys.exit(1)
