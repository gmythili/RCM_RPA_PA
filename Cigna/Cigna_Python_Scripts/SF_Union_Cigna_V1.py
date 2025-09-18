import pandas as pd
import time
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# skip_markets to skip
skip_markets = ["UL", "UM", "NE", "UG", "CW", "Other"]
#skip_markets = ["FL"]
# === File wait utility ===
def wait_for_file(filepath, timeout=20):
    print(f"Waiting for file: {filepath}")
    if os.path.basename(filepath).startswith("~$"):
        print(f"Error: {filepath} is a temporary Excel lock file. Provide actual file.")
        sys.exit(1)

    start_time = time.time()
    while True:
        try:
            if os.path.exists(filepath):
                with open(filepath, 'rb'):
                    print(f"File ready: {filepath}")
                    return
        except Exception:
            pass
        if time.time() - start_time > timeout:
            print(f"Timeout: File not accessible after {timeout} seconds -> {filepath}")
            sys.exit(1)
        time.sleep(1)

# === Startup ===
print("Starting SF_Union_Portals Cigna processing script...\n")
time.sleep(2)

# === Get args ===
if len(sys.argv) != 3:
    print("Usage: python SF_Union_Portals_Cigna.py <salesforce_excel_path> <cigna_csv_path>")
    sys.exit(1)

sf_path = sys.argv[1]
cigna_path = sys.argv[2]

wait_for_file(sf_path)
wait_for_file(cigna_path)

try:
    # === Read Cigna CSV ===
    print("Reading Cigna CSV...")
    try:
        cigna_df = pd.read_csv(cigna_path, encoding='utf-8')
    except UnicodeDecodeError:
        cigna_df = pd.read_csv(cigna_path, encoding='latin1')  # fallback

    # Clean columns
    cigna_df['EMAIL'] = cigna_df['EMAIL'].str.lower().str.strip()

    # Create lookup dict
    cigna_lookup = cigna_df.set_index('EMAIL')['STATUS'].to_dict()

    # Function to check Cigna status
    def check_cigna_status(email):
        email = str(email).strip().lower()
        status = cigna_lookup.get(email)
        if status is None:
            return 'Success - User not found'
        else:
            return email

    # === Process Salesforce Excel ===
    work_book = load_workbook(sf_path)
    all_sheets = work_book.sheetnames
    sheets_to_process = [s for s in all_sheets if s not in skip_markets]

    emailvalue_rows = []

    for sheet_name in sheets_to_process:
        print(f"Processing sheet: {sheet_name}")
        sf_df = pd.read_excel(sf_path, sheet_name=sheet_name, engine='openpyxl')

        if 'EMAIL' not in sf_df.columns:
            print(f"'EMAIL' column not found in {sheet_name}. Skipping.")
            continue

        sf_df['EMAIL'] = sf_df['EMAIL'].str.lower().str.strip()
        sf_df['CIGNA'] = sf_df['EMAIL'].apply(check_cigna_status)

        ws = work_book[sheet_name]
        ws.delete_rows(1, ws.max_row)

        for row in dataframe_to_rows(sf_df, index=False, header=True):
            ws.append(row)

        # Normalize columns for lookup
        sf_df.columns = [col.strip().upper().replace(" ", "_") for col in sf_df.columns]
        has_fname = 'FN' in sf_df.columns
        has_lname = 'LN' in sf_df.columns

        for idx, val in sf_df['CIGNA'].items():
            if isinstance(val, str) and val.endswith('.com'):
                email = sf_df.at[idx, 'EMAIL']
                first_name = sf_df.at[idx, 'FN'] if has_fname else ''
                last_name = sf_df.at[idx, 'LN'] if has_lname else ''
                row_number = idx + 2  # +2 for header
                col_number = sf_df.columns.get_loc('CIGNA') + 1
                emailvalue_rows.append([email, row_number, col_number, sheet_name, first_name, last_name])

    # Save Excel
    work_book.save(sf_path)
    print("All sheets updated successfully.")

    # === Create Output CSV ===
    if emailvalue_rows:
        timestamp = datetime.now().strftime("%d%m%Y")
        output_filename = f"{timestamp}_CignaActive.csv"
        output_path = os.path.join(os.path.dirname(sf_path), output_filename)
        emailvalue_df = pd.DataFrame(emailvalue_rows, columns=['Email', 'Row', 'Column', 'Sheet', 'First Name', 'Last Name'])
        emailvalue_df.to_csv(output_path, index=False)
        print(f"CSV file saved: {output_path}")
    else:
        print("No active Cigna emails found; no CSV created.")
    
        # === Modified block to deduplicate before saving CSV ===
    if emailvalue_rows:
        try:
            timestamp = datetime.now().strftime("%d%m%Y")
            output_path = os.path.join(os.path.dirname(sf_path), f"{timestamp}_CignaActive.csv")

            # Create DataFrame with all collected rows
            df = pd.DataFrame(emailvalue_rows, columns=['Email', 'Row', 'Column', 'Sheet', 'First Name', 'Last Name' ])

            # Remove duplicate rows based on Email, Row, Column, and Sheet
            df = df.drop_duplicates(subset=['Email', 'Row', 'Column'])

            # Write the deduplicated data to CSV
            df.to_csv(output_path, index=False)
            print(f"CSV saved: {output_path}")
        except Exception as e:
            print(f"Error writing CSV: {str(e)}")
            sys.exit(1)
    else:
        print("No active emails found. CSV not created.")

except Exception as e:
    print("An error occurred:", str(e))
    sys.exit(1)