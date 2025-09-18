import pandas as pd
import time
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Markets to skip from Cigna logic
skip_markets = ["UL", "UM", "NE", "UG", "CW", "Other"]
#markets=["FL"]
def wait_for_file(filepath, timeout=20):
    print(f"Waiting for file: {filepath}")
    if os.path.basename(filepath).startswith('~$'):
        print(f"Error: The file {filepath} looks like a temporary Excel lock file (starts with '~$').")
        sys.exit(1)

    start_time = time.time()
    while True:
        try:
            if os.path.exists(filepath):
                with open(filepath, 'rb'):
                    print(f"File ready: {filepath}")
                    return
        except Exception:
            pass
        if time.time() - start_time > timeout:
            print(f"Timeout: File not accessible after {timeout} seconds -> {filepath}")
            sys.exit(1)
        time.sleep(1)

# === Startup ===
print("Starting SF_Union Summary Only script...\n")
time.sleep(2)

# === Get command line args ===
if len(sys.argv) != 2:
    print("Usage: python SF_Union_SummaryOnly.py <salesforce_excel_path>")
    sys.exit(1)

sf_path = sys.argv[1]
wait_for_file(sf_path)

try:
    wb = load_workbook(sf_path)
    all_sheets = wb.sheetnames
    summary_data = []

    for sheet_name in all_sheets:
        if "Report" in sheet_name:  # Skip any sheet with 'Report' in its name
            continue

        df = pd.read_excel(sf_path, sheet_name=sheet_name, engine='openpyxl')
        total = len(df)

        # Default counts
        not_found = 0
        deactivate = 0
        failure = 0  # New: for Failure - Required Action

        # Skip special markets but still record total count
        if sheet_name  in skip_markets:
            summary_data.append([sheet_name, total, 0, 0, 0,])
            continue

        # Only process if 'Cigna' column is available
        if 'CIGNA' in df.columns:
            df['CIGNA'] = df['CIGNA'].astype(str).fillna("").str.strip()

            not_found = df['CIGNA'].str.contains(
                r'Success - User not found', case=False, na=False
            ).sum()

            deactivate = df['CIGNA'].str.contains(
                r'Success - User found and deactivated', case=False, na=False
            ).sum()

            # NEW: Failure - Required Action
            failure = (
                df['CIGNA']
                .str.lower()
                .str.replace('-', '', regex=False)
                .str.replace(r'\s+', ' ', regex=True)
                .str.contains("failure required action", na=False)
                .sum()
            )

        summary_data.append([sheet_name, total, not_found,deactivate, failure])

    # Remove old 'Cigna_Report' sheet if it exists
    if 'Cigna_Report' in wb.sheetnames:
        del wb['Cigna_Report']

    # Create new summary sheet
    main_ws = wb.create_sheet('Cigna_Report', 0)
    headers = [
        'Market',
        'SF_COUNT',
        'Cigna_UserNotFound',
        'Cigna_UserFoundandDeactivated',
        'Cigna_Failure'  # New header
    ]
    main_ws.append(headers)

    # Style setup
    bold_font = Font(bold=True)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in main_ws[1]:
        cell.font = bold_font
        cell.border = thin_border

    for row in summary_data:
        main_ws.append(row)

    # Add totals row
    totals = ['ALL Markets']
    for i in range(1, len(headers)):  # Sum all numeric columns
        total = sum(row[i] for row in summary_data if len(row) > i)
        totals.append(total)

    main_ws.append(totals)

    # Apply styling to totals row
    for cell in main_ws[main_ws.max_row]:
        cell.fill = green_fill
        cell.font = bold_font
        cell.border = thin_border

    # Apply borders to all other data rows
    for row in main_ws.iter_rows(min_row=2, max_row=main_ws.max_row - 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    wb.save(sf_path)
    print("Main summary sheet created successfully as the first sheet.")

except Exception as e:
    print("An error occurred:", str(e))
    sys.exit(1)